import matplotlib.pyplot as plt
import pandas as pd
from fpdf import FPDF
from datetime import datetime
import openpyxl


save_dir = "/Users/sakura/Desktop/Argus 1.21/Argus_1.21_Back_end/bg_images/datasheet/"


def addlabels(x, y, flag):
    for i in range(len(x)):
        if (y[i] >= 50):
            plt.text(i, y[i], y[i], ha='center',
                     Bbox=dict(facecolor='red', alpha=.7))
            if (flag == 0):
                plt.vlines(x[i], 1, y[i], linestyles='dashed',
                           colors='red', alpha=.2)
            plt.gca().get_xticklabels()[i].set_color("red")

        if (y[i] < 50):
            plt.text(i, y[i], y[i], ha='center',
                     Bbox=dict(facecolor='green', alpha=.7))
            if (flag == 0):
                plt.vlines(x[i], 1, y[i], linestyles='dashed',
                           colors='green', alpha=.2)
            plt.gca().get_xticklabels()[i].set_color("green")


def pdf(flag):
    # can also index sheet by name or fetch all sheets
    df = pd.read_excel(save_dir + 'demo.xlsx', engine='openpyxl')
    name = df['filename'].tolist()
    value = df['calculated percentage'].tolist()
    if (flag == 0):
        plt.plot(df["filename"], df["calculated percentage"],
                 marker='o', color='skyblue')
    if (flag == 1):
        plt.bar(df["filename"], df["calculated percentage"],
                width=0.1, align="center", color='skyblue')
    locs, labels = plt.xticks()
    plt.ylim(0, 110)
    plt.xticks(color='green', rotation=90)
    plt.title("IMAGE ANALYSIS DATA")
    plt.xlabel("Image names")
    plt.ylabel("Corrosion Percentage")
    # set(plt,'XTickLabel',r,'YColor','magenta')
    plt.tight_layout()

    addlabels(name, value, flag)

    plt.savefig(save_dir + 'graph.png')

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'I', 8)
    #pdf.cell(0, 8, 'qualiTEAS Inc.', 0, 1, 'R')
    # pdf.image('C:/Users/deepm/Desktop/qualiTEAS-Logo.png', x=180, y=1)
    pdf.set_font('Arial', 'B', 12)
    a = datetime.now()
    a = a.strftime("%m/%d/%y")
    pdf.cell(w=0, h=12, txt="DATE : "+a, ln=1)
    pdf.set_font('Arial', 'U', 12)
    pdf.cell(w=0, h=12, txt="Graph : ", ln=1)
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(w=0, h=12, txt='Number of Images : '+str(len(name)), ln=1)
    pdf.image(save_dir + 'graph.png', x=10, y=None, w=pdf.epw)
    pdf.set_font('Arial', 'U', 12)
    pdf.cell(w=0, h=12, txt="Table : ", ln=1)
    pdf.set_font('Arial', 'B''U', 12)
    pdf.cell(w=(pdf.epw/2)-60, h=8, ln=0)
    pdf.cell(w=40, h=8, txt='filename', border=1, ln=0, align='C')
    pdf.cell(w=80, h=8, txt='calculated percentage', border=1, ln=1, align='C')
    pdf.set_font('Arial', '', 10)
    for i in range(0, len(df)):
        pdf.cell(w=(pdf.epw/2)-60, h=8, ln=0)
        pdf.cell(w=40, h=8, txt=df['filename'].iloc[i],
                 border=1, ln=0, align='C')
        pdf.cell(w=80, h=8, txt=df['calculated percentage'].iloc[i].astype(
            str), border=1, ln=1, align='C')
    pdf.output(save_dir + 'report.pdf', 'F')


pdf(flag=1)
